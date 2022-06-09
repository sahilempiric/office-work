import datetime
import math
from dataclasses import dataclass

import openpyxl
import requests
from django.db.models import Count
from django.db.models.functions import TruncDate
from django.utils import timezone

from conf import OUTPUTS_DIR, WEB_HOOK_URL
from main import LOGGER
from surviral_avd.settings import SYSTEM_NO
from instagram.models import (ActionType, ActionForBotAccount,
                          ActionForOtherAccount, ActionForTargetAccount, ActionResult)
from instagram.models import (TwitterAccount)
from instagram.models import UserAvd, TweetForTargetAccount, TwitterTargetAccount
from accounts_conf import (LIKE_MIN_NUMBER_IN_A_WEEK,
        RETWEET_MIN_NUMBER_IN_A_WEEK, COMMENT_MIN_NUMBER_IN_A_WEEK)
from accounts_conf import NEW_POST_EXTERNAL_USER_NAMES


@dataclass
class Avd:
    name: str


class ActionReport:
    """Bot actions report generated from database"""

    def generate_action_stats_from_db(self, model=ActionForOtherAccount):
        LOGGER.info('Generating action report from database')
        stats_base = model.objects.filter(result=ActionResult.SUCCESSFUL
                                          ).annotate(day=TruncDate('created'))

        stats_by_group = stats_base.values(
            'owner', 'action', 'day').annotate(num=Count('action'))

        stats = {}

        total_num = {}
        for i in stats_by_group:
            day = i['day'].isoformat()
            account_id = i['owner']
            action = i['action']
            num = i['num']

            (avd, account) = self.get_owner_info(account_id)
            avd_name = avd.name
            #  item = {
            #      'AVD': avd.name,
            #      'Account': account.screen_name,
            #      'Action': self.get_action_name(action),
            #      'Date': day,
            #      'Action Count': num,
            #  }
            action_name = self.get_action_name(action)
            item = {
                'AVD': avd.name,
                'Account': account.screen_name,
                action_name: num,
            }

            # calculate total number for every action at the same day
            if day not in total_num:
                total_num[day] = {}

            if action_name not in total_num[day]:
                total_num[day][action_name] = num
            else:
                total_num[day][action_name] += num

            # Add the data 
            if day in stats:
                if avd_name in stats[day]:
                    if action_name not in stats[day][avd_name]:
                        stats[day][avd_name][action_name] = num
                    else:
                        LOGGER.warning(f'"{action_name} in the data')
                        stats[day][avd_name][action_name] += num
                else:
                    stats[day][avd_name] = item
            else:
                stats[day] = {}
                stats[day][avd_name] = item

        #  return stats
        return {'data': stats, 'total': total_num}

    def get_owner_info(self, account_id):
        account = TwitterAccount.objects.get(id=account_id)
        try:
            avd = UserAvd.objects.get(twitter_account=account)
        except:
            avd = Avd(name=f"AVD deleted of account is= {account_id}")

        return (avd, account)

    @staticmethod
    def get_action_name(action_id):
        actions = {}
        for i in ActionType.ACTION_TYPE:
            actions[i[0]] = i[1]

        return actions[action_id]

    @staticmethod
    def get_action_short_name(action_id):
        action_types = {2: "Follow",
                        3: "Like",
                        4: "Retweet",
                        5: "Comment",
                        9: "Unfollow"}
        return action_types[action_id]

    def generate_action_report_excel(self, report_file='action_report.xlsx'):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        report_file_path = OUTPUTS_DIR / report_file
        LOGGER.info(f'Generating the action report: {report_file_path}')

        models = [ActionForBotAccount, ActionForTargetAccount,
                  ActionForOtherAccount]
        for model in models:
            stats = self.generate_action_stats_from_db(model)
            data = stats['data']
            total = stats['total']
            #  LOGGER.debug(f'total: {total}')

            for date in total:
                nums = total[date]
                actions = [e for e in nums]
                if model is ActionForBotAccount:
                    account_type = 'bot'
                elif model is ActionForTargetAccount:
                    account_type = 'target'
                elif model is ActionForOtherAccount:
                    account_type = 'other'

                headers = ['AVD', 'Account']
                headers.extend(actions)

                ws = wb.create_sheet(f'{date}({account_type})')
                ws.append(headers)

                # one day data
                one_data = data[date]
                for avd in one_data:
                    raw_data = one_data[avd]
                    last_row = self.get_last_row_data(raw_data, headers)
                    ws.append(last_row)

                # add total number
                ws.append(self.get_total_row_data(total[date], headers))

        wb._sheets.sort(key=lambda ws: ws.title, reverse=True)
        if wb.sheetnames:
            wb.save(report_file_path)
        else:
            LOGGER.error('Cannot find any effective data from database')

    def get_last_row_data(self, raw_data, headers):
        last_row = []
        for k in headers:
            if k in raw_data:
                last_row.append(raw_data[k])
            else:
                last_row.append('')

        return last_row

    def get_total_row_data(self, raw_data, headers):
        last_row = []
        for k in headers:
            if k == 'AVD':
                last_row.append('')
                continue
            if k == 'Account':
                last_row.append('Total')
                continue

            if k in raw_data:
                last_row.append(raw_data[k])
            else:
                last_row.append('')

        return last_row

    @staticmethod
    def generate_action_report_text(report_date=None):
        models = [ActionForBotAccount, ActionForTargetAccount,
                  ActionForOtherAccount]
        report_dict = {}
        action_types = {2: "Follow",
                        3: "Like",
                        4: "Retweet",
                        5: "Comment",
                        9: "Unfollow"}
        LOGGER.info(f'Arranging all actions')
        date_from = datetime.datetime.now(tz=timezone.utc) - datetime.timedelta(days=1)
        LOGGER.info(f'Generating the action report on {report_date}')
        for model, model_type in zip(models, ["Bot", "Target", "Other"]):
            for action in model.objects.filter(result=ActionResult.SUCCESSFUL, created__gte=date_from):
                twitter_account = action.owner
                try:
                    avd = UserAvd.objects.get(twitter_account=twitter_account)
                    avd_name = f"AVD: {avd.name} Account: {twitter_account}"
                except:
                    avd_name = f"Deleted AVD  of twitter :{twitter_account}"
                action_name = action_types[action.action.id]
                try:
                    to_account = action.object.full_name
                except:
                    to_account = None
                if avd_name in report_dict:
                    if model_type in report_dict[avd_name]:
                        if action_name in report_dict[avd_name][model_type]:
                            report_dict[avd_name][model_type][action_name][0] += 1
                            if to_account:
                                report_dict[avd_name][model_type][action_name][1].append(to_account)
                        else:
                            report_dict[avd_name][model_type].update({action_name: [1, [to_account]]})
                    else:
                        report_dict[avd_name].update({model_type: {action_name: [1, [to_account]]}})
                else:
                    report_dict.update({avd_name: {model_type: {action_name: [1, [to_account]]}}})
        total_report = {}
        text = ""
        for avd_name, model_types in report_dict.items():
            text += "*" * 100 + "\n"
            text += avd_name + "\n"
            text += "*" * 100 + "\n"
            to_accounts = []
            for model_type in model_types:
                actions_dict = {}
                for action_type, action_name in action_types.items():
                    if action_name in model_types[model_type]:
                        if action_name in total_report:
                            total_report[action_name] += model_types[model_type][action_name][0]
                        else:
                            total_report[action_name] = model_types[model_type][action_name][0]
                        actions_dict[action_name] = model_types[model_type][action_name][0]
                        if model_types[model_type][action_name][1]:
                            to_accounts.extend(model_types[model_type][action_name][1])
                    else:
                        actions_dict[action_name] = 0
                text += f"{model_type} {set(to_accounts)}: {actions_dict}\n"
            text += "\n"

        # Total report
        text += "*" * 100 + "\n"
        text += "Total:\n"
        text += "*" * 100 + "\n"
        for action_name, count in total_report.items():
            text += f"{action_name}: {count}\n"
        text += "*" * 100 + "\n"
        return text

    @staticmethod
    def generate_total_report_text(report_date=None, post_report=False):
        total_actions = 0
        text = ""
        models = [ActionForBotAccount, ActionForTargetAccount,
                  ActionForOtherAccount]
        report_dict = {}
        action_types = {3: "Likes",
                        2: "Follows",
                        4: "Retweets",
                        5: "Comments"}
        LOGGER.info(f'Arranging all actions')
        date_format = '%Y-%m-%d %Z'
        date_from = datetime.datetime.strptime(f"{report_date} UTC", date_format)
        LOGGER.info(f'Generating the action report on {date_from.date()}')
        date_from = timezone.utc.localize(date_from)
        date_to = date_from + datetime.timedelta(days=1)
        print(date_from, date_to)

        active_accounts = TwitterAccount.objects.filter(status="ACTIVE")
        heading_text = f"PC {SYSTEM_NO} ({active_accounts.count()} Accounts) Date:{date_from}\n\n"
        action_text = ""
        for action in action_types:
            action_text += f"{action_types[action]} / Total\n"
            for action_model, model_type in zip(models, ["Bot", "Target", "Other"]):
                model = action_model.objects.filter(result=ActionResult.SUCCESSFUL, created__gte=date_from,
                                                    created__lte=date_to, action=action)
                action_text += f"- {model_type}: {model.count()}\n"
                total_actions += model.count()
            action_text += "\n"
        text += heading_text
        text += action_text
        payload_text = ""

        date_from = datetime.datetime.now() - datetime.timedelta(days=1)
        date_from = timezone.utc.localize(date_from)
        accounts_created_in_last_24 = TwitterAccount.objects.filter(created__gte=date_from).count()
        accounts_updated_in_last_24 = TwitterAccount.objects.filter(updated__gte=date_from,
                                                                    profile_updated=True).count()
        payload_text += heading_text
        if total_actions:
            payload_text += action_text
        accounts_created_text = f"Total accounts created in last 24 hours: {accounts_created_in_last_24}\n"
        profile_updated_text = f"Total accounts updated in last 24 hours: {accounts_updated_in_last_24}\n"
        text += accounts_created_text
        text += profile_updated_text
        if accounts_created_in_last_24:
            payload_text += accounts_created_text
        if accounts_updated_in_last_24:
            payload_text += profile_updated_text
        if post_report:
            if total_actions or accounts_created_in_last_24 or accounts_updated_in_last_24:
                payload = {
                    "text": payload_text
                }
                r = requests.post(WEB_HOOK_URL, json=payload)
                text += f"WEB HOOK Post Response:\n"
                text += f"{r.text}"
                LOGGER.info(f"WEB HOOK Post Response:")
                LOGGER.info(r.text)
        return text

    @staticmethod
    def genearte_text_report_for_one_day_total(model, total):
        action_total_stats = {}
        for action in total:
            action_total_stats[action] = {}
            action_total_stats[action][model.__name__] = total[action]
        #  LOGGER.debug(f'action_total_stats: {action_total_stats}')
        return action_total_stats

    @staticmethod
    def create_text_for_total(total_stats):
        text = ''
        for action in total_stats:
            action_text = f'{action} / Total\n'
            account_stats = total_stats[action]
            for account_type in account_stats:
                if 'ActionForTargetAccount' == account_type:
                    bot_type = 'Target'
                elif 'ActionForOtherAccount' == account_type:
                    bot_type = 'Other'
                elif 'ActionForBotAccount' == account_type:
                    bot_type = 'Bot'

                account_text = f'- {bot_type}: {account_stats[account_type]}\n'
                action_text += account_text
            text += f'{action_text}\n'

        return text

    @staticmethod
    def genearte_text_report_for_one_day_data(model, data):
        action_data_stats = {}
        for name in data:
            stats_info = data[name]
            avd_name = stats_info.pop('AVD')
            account = stats_info.pop('Account')
            actions = stats_info

            action_data_stats[name] = {}
            for action in actions:
                if action not in action_data_stats[name]:
                    action_data_stats[name][action] = {
                        model.__name__: actions[action]}
                else:
                    action_data_stats[name][action].update(
                        {model.__name__: actions[action]})
        #  LOGGER.debug(f'data: {data}')
        #  LOGGER.debug(f'action_data_stats: {action_data_stats}')
        return action_data_stats

    @staticmethod
    def create_text_for_data(data_stats):
        text = ''
        for name in data_stats:
            actions = data_stats[name]
            for action in actions:
                action_text = f'{action} / {name}\n'
                account_stats = actions[action]
                for account_type in account_stats:
                    if 'ActionForTargetAccount' == account_type:
                        bot_type = 'Target'
                    elif 'ActionForOtherAccount' == account_type:
                        bot_type = 'Other'
                    elif 'ActionForBotAccount' == account_type:
                        bot_type = 'Bot'

                    account_text = f'- {bot_type}: {account_stats[account_type]}\n'
                    action_text += account_text
                text += f'{action_text}\n'

        return text

def get_actions_for_target(start_time, end_time, action_type=None,
        include_endtime=False, target_user='', tweet_content=''):
    # get tweet record
    if target_user and tweet_content:
        tweet_owner= TwitterTargetAccount.objects.get(
                screen_name=target_user)
        tweet_record = TweetForTargetAccount.objects.get(
                owner=tweet_owner, text=tweet_content)
    else:
        tweet_record = None

    if action_type:
        if include_endtime:
            if tweet_record:
                actions = ActionForTargetAccount.objects.filter(
                        result=ActionResult.SUCCESSFUL,
                        action=action_type,
                        created__gte=start_time,
                        created__lte=end_time,
                        tweet=tweet_record)
            else:
                actions = ActionForTargetAccount.objects.filter(
                        result=ActionResult.SUCCESSFUL,
                        action=action_type,
                        created__gte=start_time,
                        created__lte=end_time)
        else:
            if tweet_record:
                actions = ActionForTargetAccount.objects.filter(
                        result=ActionResult.SUCCESSFUL,
                        action=action_type,
                        created__gte=start_time,
                        created__lt=end_time,
                        tweet=tweet_record)
            else:
                actions = ActionForTargetAccount.objects.filter(
                        result=ActionResult.SUCCESSFUL,
                        action=action_type,
                        created__gte=start_time,
                        created__lt=end_time)
    else:
        if include_endtime:
            if tweet_record:
                actions = ActionForTargetAccount.objects.filter(
                        result=ActionResult.SUCCESSFUL,
                        created__gte=start_time,
                        created__lte=end_time,
                        tweet=tweet_record)
            else:
                actions = ActionForTargetAccount.objects.filter(
                        result=ActionResult.SUCCESSFUL,
                        created__gte=start_time,
                        created__lte=end_time)
        else:
            if tweet_record:
                actions = ActionForTargetAccount.objects.filter(
                        result=ActionResult.SUCCESSFUL,
                        created__gte=start_time,
                        created__lt=end_time,
                        tweet=tweet_record)
            else:
                actions = ActionForTargetAccount.objects.filter(
                        result=ActionResult.SUCCESSFUL,
                        created__gte=start_time,
                        created__lt=end_time)
    return actions

def get_this_week_start_end_day():
    now = timezone.now()
    start_daytime = now - timezone.timedelta(days=now.weekday())
    real_start_daytime = timezone.datetime(
            start_daytime.year, start_daytime.month, start_daytime.day,
            tzinfo=now.tzinfo)
    this_week_start_day = real_start_daytime
    return (this_week_start_day, now)

def get_this_week_total_actions(target_user='', tweet_content=''):
    start_day, end_day = get_this_week_start_end_day()
    this_week_actions = get_actions_for_target(start_day, end_day,
            include_endtime=True, target_user=target_user,
            tweet_content=tweet_content)
    LOGGER.info(f'Start day: {start_day}, End day: {end_day}')
    LOGGER.info(f'Number of all actions this week: {this_week_actions.count()}')
    return this_week_actions

def get_this_week_specific_action(action_type=ActionType.LIKE_TWEET,
        target_user='', tweet_content=''):
    start_day, end_day = get_this_week_start_end_day()
    this_week_actions = get_actions_for_target(start_day, end_day,
            include_endtime=True, action_type=action_type,
            target_user=target_user, tweet_content=tweet_content)
    LOGGER.info(f'Start day: {start_day}, End day: {end_day}')
    LOGGER.info(f'Target user: {target_user}, tweet: {tweet_content}')
    LOGGER.info(f'Number of action "{action_type}" this week:'
            f' {this_week_actions.count()}')
    return this_week_actions

def get_this_week_like_action(target_user='', tweet_content=''):
    action_type = ActionType.LIKE_TWEET
    return get_this_week_specific_action(action_type=action_type,
            target_user=target_user, tweet_content=tweet_content)

def get_this_week_retweet_action(target_user='', tweet_content=''):
    action_type = ActionType.RETWEET_TWEET
    return get_this_week_specific_action(action_type=action_type,
            target_user=target_user, tweet_content=tweet_content)

def get_this_week_comment_action(target_user='', tweet_content=''):
    action_type = ActionType.COMMENT_TWEET
    return get_this_week_specific_action(action_type=action_type,
            target_user=target_user, tweet_content=tweet_content)

def get_this_week_like_action_number(target_user='', tweet_content=''):
    return get_this_week_like_action(
            target_user=target_user, tweet_content=tweet_content).count()

def get_this_week_retweet_action_number(target_user='', tweet_content=''):
    return get_this_week_retweet_action(
            target_user=target_user, tweet_content=tweet_content).count()

def get_this_week_comment_action_number(target_user='', tweet_content=''):
    return get_this_week_comment_action(
            target_user=target_user, tweet_content=tweet_content).count()

def get_target_account_from_tweet(tweet_id):
    return TweetForTargetAccount.objects.get(id=tweet_id)

def get_target_user_from_tweet(tweet_id):
    return TweetForTargetAccount.objects.get(id=tweet_id).owner.screen_name

def get_this_week_statistics_of_actions(target_user='', tweet_content=''):
    actions = get_this_week_total_actions(target_user=target_user,
            tweet_content=tweet_content)
    statistics = actions.exclude(tweet=None).values(
            'tweet', 'action').annotate(Count('tweet'))
    stats = []
    for r in statistics:
        tweet_id = r['tweet']
        target_user = get_target_user_from_tweet(tweet_id)
        #  action = ActionReport.get_action_name(r['action'])
        action = ActionReport.get_action_short_name(r['action'])
        count = r['tweet__count']
        stats.append(
                {'target_user': target_user, 'tweet_id': tweet_id,
                    'action': action, 'count': count})
    return stats

def print_this_week_stats(target_user='', tweet_content=''):
    stats = get_this_week_statistics_of_actions(target_user=target_user,
            tweet_content=tweet_content)
    print('Engagement number for actions like, retweet, and comment:')
    for r in stats:
        print(f'Target user: {r["target_user"]:<20}Tweet id: {r["tweet_id"]}\t'
                f'Action: {r["action"]:<10}Count: {r["count"]}')

def check_comment_number_this_week(target_user='', tweet_content=''):
    db_number = get_this_week_comment_action_number(
            target_user=target_user, tweet_content=tweet_content)

    LOGGER.info(f'Target user: {target_user},'
            f' Comment number in DB: {db_number},'
            f' Min number: {COMMENT_MIN_NUMBER_IN_A_WEEK}')
    if db_number < COMMENT_MIN_NUMBER_IN_A_WEEK:
        LOGGER.debug('The comment number in DB is less then the min number')
        return False
    else:
        LOGGER.debug('The comment number in DB is greater then the min number')
        return True

def check_like_number_this_week(target_user='', tweet_content=''):
    db_number = get_this_week_like_action_number(
            target_user=target_user, tweet_content=tweet_content)

    LOGGER.info(f'Target user: {target_user},'
            f' Like number in DB: {db_number},'
            f' Min number: {LIKE_MIN_NUMBER_IN_A_WEEK}')
    if db_number < LIKE_MIN_NUMBER_IN_A_WEEK:
        LOGGER.debug('The like number in DB is less then the min number')
        return False
    else:
        LOGGER.debug('The like number in DB is greater then the min number')
        return True

def check_retweet_number_this_week(target_user='', tweet_content=''):
    db_number = get_this_week_retweet_action_number(
            target_user=target_user, tweet_content=tweet_content)

    LOGGER.info(f'Target user: {target_user},'
            f' Retweet number in DB: {db_number},'
            f' Min number: {RETWEET_MIN_NUMBER_IN_A_WEEK}')
    if db_number < RETWEET_MIN_NUMBER_IN_A_WEEK:
        LOGGER.debug('The retweet number in DB is less then the min number')
        return False
    else:
        LOGGER.debug('The retweet number in DB is greater then the min number')
        return True

def get_like_quota_one_day(period=7):
    quota = math.ceil(LIKE_MIN_NUMBER_IN_A_WEEK / period)
    return quota

def get_retweet_quota_one_day(period=7):
    quota = math.ceil(RETWEET_MIN_NUMBER_IN_A_WEEK / period)
    return quota

def get_comment_quota_one_day(period=7):
    quota = math.ceil(COMMENT_MIN_NUMBER_IN_A_WEEK / period)
    return quota

def get_quotas_until_today_for_action(action, period=7):
    today = timezone.now().date()
    days = today.isoweekday()
    
    if action == 'like':
        quota = get_like_quota_one_day(period=period)
    elif action == 'retweet':
        quota = get_retweet_quota_one_day(period=period)
    elif action == 'comment':
        quota = get_comment_quota_one_day(period=period)

    quotas = quota * days
    LOGGER.debug(f'Days: {days}, Action: {action}, Quotas: {quotas}')
    return quotas

def get_target_names_standard_and_nonstandard_for_one_day():
    all_target_names = NEW_POST_EXTERNAL_USER_NAMES
    standard_target_names = set()
    nonstandard_target_names = set()
    stats = get_this_week_statistics_of_actions()
    #
    #  like_quota = get_like_quota_one_day()
    #  retweet_quota = get_retweet_quota_one_day()
    #  comment_quota = get_comment_quota_one_day()

    like_quota = get_quotas_until_today_for_action(action='like')
    retweet_quota = get_quotas_until_today_for_action(action='retweet')
    comment_quota = get_quotas_until_today_for_action(action='comment')

    LOGGER.debug(f'Like quota: {like_quota}, Retweet quota: {retweet_quota}, '
            f'Comment quota: {comment_quota}')

    for r in stats:
        if 'Like' in r['action']:
            if r['count'] < like_quota:
                nonstandard_target_names.add(r['target_user'])
            else:
                standard_target_names.add(r['target_user'])

        if 'Retweet' in r['action']:
            if r['count'] < retweet_quota:
                nonstandard_target_names.add(r['target_user'])
            else:
                standard_target_names.add(r['target_user'])

        if 'Comment' in r['action']:
            if r['count'] < comment_quota:
                nonstandard_target_names.add(r['target_user'])
            else:
                standard_target_names.add(r['target_user'])

    # if any action of one target user cannot reach the quota,
    # although some actions reache the quota,
    # then add it to the real_nonstandard_target_names
    real_standard_target_names = {e for e in standard_target_names
            if e not in standard_target_names}
    real_nonstandard_target_names = {e for e in all_target_names
            if e not in standard_target_names}
    LOGGER.debug(f'Standard target users: {real_standard_target_names}')
    LOGGER.debug(f'Nonstandard target users: {real_nonstandard_target_names}')

    return real_nonstandard_target_names, real_standard_target_names

